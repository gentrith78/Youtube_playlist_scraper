import googleapiclient.discovery
from openpyxl import  load_workbook
import requests
from playwright.async_api import async_playwright

import asyncio
from urllib.parse import parse_qs, urlparse

import credentials

# dir_to_save_video = 'C:\\Users\\38343\\Desktop\\image_to_convert\\'
dir_to_save_video = 'Downloaded_Songs\\'


#extract playlist id from url


playlist_url = 'playlist url'
query = parse_qs(urlparse(playlist_url).query, keep_blank_values=True)
playlist_id = query["list"][0]
print(f'get all playlist items links from {playlist_id}')
youtube = googleapiclient.discovery.build("youtube", "v3", developerKey = credentials.api_key)
#extract data for each video on playlist
request = youtube.playlistItems().list(
    part = "snippet",
    playlistId = playlist_id,
    maxResults = 50
)
response = request.execute()

playlist_items = []
while request is not None:
    response = request.execute()
    playlist_items += response["items"]
    request = youtube.playlistItems().list_next(request, response)

def get_video_duration(id):
    res = requests.get(f'https://www.googleapis.com/youtube/v3/videos?part=contentDetails&id={id}&key={credentials.api_key}')
    data = res.json()['items'][0]['contentDetails']['duration'][2:-1].split(sep='M')
    return f'{data[0]} Min--{data[1]} Secs'


def get_video_data(id):
    try:
        res = requests.get(f'https://www.googleapis.com/youtube/v3/videos?part=statistics&id={id}&key={credentials.api_key}')
        data = res.json()['items'][0]['statistics']
    except:
        data = ''
    return data

playlist_data=[]
print('Extracting Video Data...')
for el in playlist_items:
    stats = get_video_data(el["snippet"]["resourceId"]["videoId"])
    try:
        views = stats['viewCount']
    except:
        views = ''
    try:
        coments = stats['commentCount']
    except:
        coments = ''
    try:
        likes = stats['likeCount']
    except:
        likes = ''
    try:
        duration = get_video_duration(el["snippet"]["resourceId"]["videoId"])
    except:
        duration = ''
    try:
        data = {'publishment':el['snippet']['publishedAt'],
                'title':el['snippet']['title'],
                # 'description':el['snippet']['description'],
                'thumbnail':el['snippet']['thumbnails']['default']['url'],
                'link':f'https://www.youtube.com/watch?v={el["snippet"]["resourceId"]["videoId"]}&list={playlist_id}&t=0s',
                'views':views,
                'likes':likes,
                'coments':coments,
                'duration':duration
                }
        playlist_data.append(data)
        print(data['title'])
    except:
        continue


#download video with online converter through web automation with playwright
links = [{'link':i['link'],'title':i['title']} for i in playlist_data]
print(f'Starting Download Procces ({len(links)} videos )...')
iteration = len(links)
async def run(browser):
    downloaded_content = []
    context = await browser.new_context(viewport={ 'width': 300, 'height': 400 })
    page = await context.new_page()
    global iteration
    for _ in range(100):
        try:
            if len(links) > 0:
                url = links[0]['link']
                title = links[0]['title']
                links.pop(0)
            else:break
            await page.goto('https://ytmp3.cc/uu129cc/')

            #fill link
            await page.type('//*[@id="input"]',url,delay=10)
            #convert
            await page.click('//*[@id="submit"]')
            #download
            try:
                await page.wait_for_selector('//*[@id="download"]',timeout=10000)
            except:
                iteration -= 1
                print(f'Could not download1 {title}')
                continue
            download_button = await page.query_selector('//*[@id="download"]')
            context.on('page', lambda page: page.close())
            try:
                async with page.expect_download() as download_info:
                    try:
                        page.set_default_timeout(10000)
                        await download_button.click()
                    except:
                        iteration -= 1
                        print(f'Could not download2 {title}')
                        continue
                    try:
                        await page.wait_for_selector('//*[@id="error"]',timeout=300)
                        iteration -= 1
                        print(f'Could not download3 {title}')
                        continue
                    except:
                        pass
            except:
                print(f'Could not download3.5 {title}')
                continue
            download = await download_info.value
            downloaded_content.append(download)
            iteration -= 1
            print(f'DOWNLOADING --------{iteration}------- {download.suggested_filename} ')
        except:
            print(f'Could not download3.5 {title}')
            continue

    for el in downloaded_content:
        try:
            await el.save_as(f'{dir_to_save_video}{el.suggested_filename}')
            print(f"----------------------DOWNLOAD FINISHED {el.suggested_filename}")
        except:
            iteration -= 1
            print(f'cold not download5 {el.suggested_filename}')

async def main():
    async with async_playwright() as playwright:
        tasks = []
        chromium = playwright.chromium
        browser = await chromium.launch(headless=False)
        for i in range(3):
            task = asyncio.ensure_future(run(browser=browser))
            tasks.append(task)

        await asyncio.gather(*tasks)

loop = asyncio.get_event_loop()
loop.run_until_complete(main())

#store data to excel
print('Puting Data Into Excel...')

wb = load_workbook('Playlist__data.xlsx')

worksheet = wb['Data']
worksheet_stats = wb['Stats']

#enter data in to cells
for i in range(len(playlist_data)):
    worksheet.cell(row=i+2, column=1).value = playlist_data[i]['title']
    worksheet.cell(row=i+2, column=2).value = playlist_data[i]['publishment']
    worksheet.cell(row=i+2, column=3).value = playlist_data[i]['duration']
    worksheet.cell(row=i+2, column=4).value = playlist_data[i]['link']
    worksheet.cell(row=i+2, column=5).value = playlist_data[i]['views']
    worksheet.cell(row=i+2, column=6).value = playlist_data[i]['coments']
    worksheet.cell(row=i+2, column=7).value = playlist_data[i]['likes']
    worksheet.cell(row=i+2, column=8).value = playlist_data[i]['thumbnail']

#Getting Statistics (Most Viewed Video, Most Liked Viedo
#sort most viewed videos
#Get the views
views = list(int(i['views']) for i in playlist_data)
views.sort(reverse=True)
#algorithm to write into cells
cell_nr = 2
for view in views:
    for el in playlist_data:
        if int(el['views']) == view:
            worksheet_stats.cell(row=cell_nr, column=2).value = view
            worksheet_stats.cell(row=cell_nr, column=1).value = el['title']
            cell_nr+=1



wb.save('Playlist__data.xlsx')



